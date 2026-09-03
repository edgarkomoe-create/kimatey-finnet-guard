"""
Export enrichi des resultats d'analyse - reutilise par les 3 modules
(Reseau, Transactions, Securite IIoT).

Repond a un probleme reel et documente en cybersecurite : la "fatigue d'alerte"
(alert fatigue) - une liste plate de resultats, dans l'ordre d'origine, oblige
l'analyste a tout trier/filtrer manuellement pour savoir quoi traiter en
premier. Cet export resout ca par un "harnais ajustable" (meme principe que
core/schema_router.py et la strategie de desequilibre du pipeline IIoT) :

1. TOUJOURS applique (deterministe, instantane, gratuit, sur toutes les lignes) :
   - Une colonne Priorite (gravite x confiance)
   - Un tri automatique (plus urgent en premier)
   - Une mise en forme conditionnelle (couleurs) dans le fichier Excel

2. REGLABLE par l'utilisateur (IA, plus riche mais plus lente/couteuse) :
   - Aucune recommandation IA (par defaut, rapide)
   - Recommandation Lieutenant Cyber sur le Top 5 ou Top 10 des cas les plus
     prioritaires uniquement - jamais sur l'integralite du fichier (inutile
     pour les cas mineurs, et beaucoup trop lent/couteux a grande echelle).
"""
import io

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Poids de gravite par defaut - meme echelle que core/alert_log.py (1=faible,
# 2=eleve, 3=critique) pour rester coherent avec le score de securite deja
# utilise ailleurs dans le systeme.
COULEUR_CRITIQUE = "F8696B"   # rouge
COULEUR_ELEVEE = "FFC000"     # orange
COULEUR_MOYENNE = "FFEB84"    # jaune
COULEUR_FAIBLE = "C6E0B4"     # vert clair


def calculer_priorite(df: pd.DataFrame, col_classe: str, col_confiance: str, poids_gravite: dict) -> pd.DataFrame:
    """Ajoute une colonne 'Priorite' (score numerique) et 'Niveau_Priorite'
    (texte) a partir de la gravite de la classe predite et de la confiance.
    Formule volontairement simple et transparente (gravite x confiance/100) -
    pas une boite noire, un score explicable a un analyste ou un jury."""
    df = df.copy()
    gravite = df[col_classe].map(poids_gravite).fillna(0)
    confiance_normalisee = df[col_confiance].astype(float) / 100
    df["Priorite"] = (gravite * confiance_normalisee).round(3)

    def niveau(score):
        if score >= 2.4:
            return "🔴 Critique"
        elif score >= 1.5:
            return "🟠 Elevee"
        elif score >= 0.5:
            return "🟡 Moyenne"
        else:
            return "🟢 Faible"

    df["Niveau_Priorite"] = df["Priorite"].apply(niveau)
    return df.sort_values("Priorite", ascending=False).reset_index(drop=True)


def generer_recommandations_ia(df_top: pd.DataFrame, col_classe: str, col_confiance: str,
                                ask_gemini_fn, client, system_prompt) -> list:
    """Genere une courte recommandation Lieutenant Cyber pour chaque ligne du
    Top N fourni (jamais sur l'integralite du fichier - voir docstring module).
    Retourne une liste de chaines, une par ligne, dans le meme ordre que df_top.
    En cas d'echec (pas de cle Gemini, erreur API), retourne une chaine vide
    plutot que de faire planter tout l'export."""
    recommandations = []
    for _, row in df_top.iterrows():
        try:
            contenu = f"Categorie detectee : {row[col_classe]}, confiance : {row[col_confiance]}%."
            avis = ask_gemini_fn(client, contenu, system_instruction=system_prompt)
            recommandations.append(avis)
        except Exception:
            recommandations.append("")
    return recommandations


def construire_excel_enrichi(df: pd.DataFrame, col_classe: str, col_confiance: str,
                              poids_gravite: dict, titre_feuille_detail: str = "Detail",
                              recommandations_top_n: dict = None) -> bytes:
    """Construit un fichier Excel a 2 feuilles (Resume + Detail), avec priorite,
    tri et coloration conditionnelle. recommandations_top_n, si fourni, est un
    dict {index_ligne: texte_recommandation} pour les lignes les plus
    prioritaires - ajoute une colonne 'Recommandation IA' uniquement pour ces
    lignes-la, vide pour les autres (jamais calcule pour tout le fichier)."""
    df_priorise = calculer_priorite(df, col_classe, col_confiance, poids_gravite)

    if recommandations_top_n:
        df_priorise["Recommandation_IA"] = df_priorise.index.map(lambda i: recommandations_top_n.get(i, ""))

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # ---- Feuille Resume ----
        resume_counts = df_priorise[col_classe].value_counts().reset_index()
        resume_counts.columns = ["Categorie", "Nombre"]
        resume_counts.to_excel(writer, sheet_name="Resume", index=False, startrow=1)

        top10 = df_priorise.head(10)[[col_classe, col_confiance, "Priorite", "Niveau_Priorite"]]
        top10.to_excel(writer, sheet_name="Resume", index=False, startrow=len(resume_counts) + 5)

        # ---- Feuille Detail (triee, coloree) ----
        df_priorise.to_excel(writer, sheet_name=titre_feuille_detail, index=False)

        ws = writer.sheets[titre_feuille_detail]
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value=f"Resultats tries par priorite - {len(df_priorise)} lignes analysees")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)

        # En-tetes en gras
        header_row = 2
        for col_idx in range(1, len(df_priorise.columns) + 1):
            ws.cell(row=header_row, column=col_idx).font = Font(bold=True)

        # Coloration conditionnelle par ligne selon Niveau_Priorite
        try:
            col_niveau_idx = list(df_priorise.columns).index("Niveau_Priorite") + 1
        except ValueError:
            col_niveau_idx = None

        couleur_par_niveau = {
            "🔴 Critique": COULEUR_CRITIQUE, "🟠 Elevee": COULEUR_ELEVEE,
            "🟡 Moyenne": COULEUR_MOYENNE, "🟢 Faible": COULEUR_FAIBLE,
        }
        if col_niveau_idx:
            for row_idx, niveau_val in enumerate(df_priorise["Niveau_Priorite"], start=header_row + 1):
                couleur = couleur_par_niveau.get(niveau_val)
                if couleur:
                    fill = PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")
                    for col_idx in range(1, len(df_priorise.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

        # Largeur de colonnes raisonnable
        for col_idx, col_name in enumerate(df_priorise.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(35, max(12, len(str(col_name)) + 4))

    buffer.seek(0)
    return buffer.getvalue()
